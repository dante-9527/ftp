import os
from openpyxl import load_workbook, workbook
import config
from utils.commons import *


class ViewInterfaces(object):

    def __init__(self, conn):
        self.conn = conn
        self.is_login = None

    @staticmethod
    def save_userinfo(user_dic, path=config.USER_INFO_PATH):
        """储存用户信息"""
        # 判断用户excel是否存在，若不存在则创建文件，存在则追加写入
        if not os.path.exists(path):
            head_info = ["用户名", "密码", "注册时间"]
            wb = workbook.Workbook()
            sheet = wb.worksheets[0]
            for index, head in enumerate(head_info, 1):
                cell = sheet.cell(1, index)
                cell.value = head
        else:
            wb = load_workbook(path)
            sheet = wb.worksheets[0]
            # 判断用户是否已注册
            for row in sheet.rows:
                if user_dic['username'] == row[0].value:
                    # 若存在则返回(False, "用户名已存在>>>")
                    return False, "用户名已存在>>>"
        # 获取行数最大值
        mrows = sheet.max_row
        for index, data in enumerate(user_dic.values(), 1):
            cell = sheet.cell(mrows + 1, index)
            cell.value = data
        wb.save(path)
        # 建立用户文件夹
        os.makedirs(os.path.join(config.USER_DATA_DIR, user_dic['username']))
        # 建好之后返回(True, "注册成功>>>")
        return True, "注册成功>>>"

    def register_interface(self):
        """注册接口"""
        print("注册功能开启>>>")
        user_dic = recv_data(self.conn)
        # 客户端退出注册传输'False'，服务端也退出注册
        if not user_dic:
            return
        reply = self.save_userinfo(user_dic)
        send_data(self.conn, reply)

    def login(self, user_dic, path=config.USER_INFO_PATH):
        """登录功能"""
        # 判断用户是否注册
        if not os.path.exists(path):
            # 用户未注册则返回(False, "用户不存在，请先注册>>>")
            return False, "用户不存在，请先注册>>>"
        wb = load_workbook(path)
        sheet = wb.worksheets[0]
        for row in sheet.rows:
            if user_dic['username'] == row[0].value and user_dic['pwd'] == row[1].value:
                # 用户登录成功则修改登录状态
                self.is_login = user_dic['username']
                return True, "登录成功>>>"
        return False, "用户名或密码错误>>>"

    def login_interface(self):
        """登录接口"""
        print("登录功能开启>>>")
        user_dic = recv_data(self.conn)
        if not user_dic:
            return
        reply = self.login(user_dic)
        send_data(self.conn, reply)

    def check_file(self):
        """查询接口"""
        # 获取登录状态
        file_list = os.listdir(os.path.join(config.USER_DATA_DIR, self.is_login))
        if not file_list:
            return False, "您的网盘上没有文件>>>"
        return True, file_list

    def check_file_interface(self):
        if not self.is_login:
            print("用户未登录>>>")
        print("查询功能开启>>>")
        reply = self.check_file()
        send_data(self.conn, reply)

    def recv_file(self, save_file_name, chunk_size=1024):
        save_file_path = os.path.join(config.USER_DATA_DIR, self.is_login, save_file_name)
        # 获取头部信息，数据长度
        has_read_size = 0
        bytes_list = []
        while has_read_size < 4:
            chunk = self.conn.recv(4 - has_read_size)
            bytes_list.append(chunk)
            has_read_size += len(chunk)
        header = b"".join(bytes_list)
        data_length = struct.unpack('i', header)[0]

        # 获取数据
        with open(save_file_path, mode="wb") as file_object:
            has_write_data_size = 0
            while has_write_data_size < data_length:
                size = chunk_size if (data_length - has_write_data_size) > chunk_size \
                    else (data_length - has_write_data_size)
                chunk = self.conn.recv(size)
                file_object.write(chunk)
                file_object.flush()
                has_write_data_size += len(chunk)
                print_progress_bar(has_write_data_size, data_length)
            print("\n文件上传成功>>>")
        return True, "文件上传成功>>>"

    def send_file(self, file_path, c_file_size):
        # 计算需要传送的文件大小
        file_size = os.stat(file_path).st_size  # 文件总大小
        send_size = file_size - c_file_size  # 需要传输的大小
        header = struct.pack('i', send_size)
        self.conn.sendall(header)

        has_send_size = c_file_size  # 已经传输的大小
        need_send_size = 0  # 后续传输的大小
        file_object = open(file_path, mode='rb')
        # 将指针移动到对应位置
        file_object.seek(c_file_size, 0)

        while need_send_size < send_size:
            chunk = file_object.read(2048)
            try:
                self.conn.sendall(chunk)
                need_send_size += len(chunk)
                has_send_size += len(chunk)
                print_progress_bar(has_send_size, file_size)
            except BlockingIOError:
                if need_send_size == send_size:
                    break
                continue
        file_object.close()

    def upload_interface(self):
        print(self.is_login)
        if not self.is_login:
            return
        print("上传功能开启>>>")
        save_file_name = recv_data(self.conn)
        if not save_file_name:
            return
        reply = self.recv_file(save_file_name)
        send_data(self.conn, reply)

    def download_interface(self):
        if not self.is_login:
            return
        print("下载功能开启>>>")
        # 接收文件信息
        file_info = recv_data(self.conn)
        if not file_info:
            return
        # print(file_info)
        c_file_name = file_info.get('file_name')
        c_file_size = file_info.get('file_size')
        file_path = os.path.join(config.USER_DATA_DIR, self.is_login, c_file_name)
        # 判断文件是否存在
        if not os.path.exists(file_path):
            # 传输给客户端反馈信息flag
            send_data(self.conn, False)
            return
        # 若存在则传输flag
        send_data(self.conn, True)
        # 计算服务端文件大小
        file_size = os.stat(file_path).st_size
        # 如果客户端文件大于服务端文件，则将c_file_size改成0，重新下载
        if c_file_size > file_size:
            c_file_size = 0
        # 将文件大小传输给客户端
        send_data(self.conn, file_size)
        # 接收客户端穿回来的下载指令
        choice = recv_data(self.conn)
        # 若客户想重新下载，则将c_file_size改成0
        if choice.upper() == "N":
            c_file_size = 0
        self.send_file(file_path, c_file_size)
        print("\n传输完成>>>")
